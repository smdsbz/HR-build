#!/usr/bin/python3
# -*- coding: utf-8 -*-

'   self-made utils for operating .xlsx files based on openpyxl ---- by smdsbz   '

from openpyxl import Workbook, load_workbook
from datetime import datetime
import os, sqlite3

######## general config ########

FOLDER = os.path.join(os.curdir, 'score-sheets')  # xlsx location
DATABASE = os.path.join(os.curdir, 'data.db')  # db loaction
# that is to say: main.py, xlsxSwissKnife.py and data.da must be under the same dir

######## utils ########

def _read_test():
	''' return table header line '''
	wb = load_workbook(os.path.join(FOLDER, 'template.xlsx'))
	data = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	return tuple([ content.value for content in data['A2':'M2'][0] ])

def _move_cursor(ws, name='something you have to mess up with'):
	'''
	return the index asked
	IF no match THEN return index-number of the adjacent empty row
	'''
	nameCells = ws['A'][2:] # 1: test-use; running: 2
	names = tuple(filter(lambda s: s and s.strip(), [ content.value for content in nameCells ]))  # thx 2 MichealLiao
	if name in names:
		return names.index(name) + 3 # 2: test-use
	else:
		return len(names) + 3

def newFile(title="测试测试", depart="其它", *, date=str(datetime.now())):
	''' derive a new .xlsx from ./score-sheets/template.xlsx '''
	filename = title + ' - ' + date + '.xlsx'
	dst = os.path.join(FOLDER, filename)
	try:
		wb = load_workbook(os.path.join(FOLDER, 'template.xlsx'))
	except IOError:
		return 0
	else:
		ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
		ws.title = title
		ws['B1'].value, ws['F1'].value, ws['J1'].value = title, depart, date
		with sqlite3.connect(DATABASE) as database:
			cursor = database.execute("select name from test where depart = '%s'" % depart)
			names = cursor.fetchall()
		for i in range(len(names)):
			ws['A'+str(i+3)].value = names[i][0]
		wb.save(dst)
		return 1

def write(filname, data_in):
	''' write/update one persona at a time '''
	dst = os.path.join(FOLDER, filename)
	try:
		wb = load_workbook(dst)
	except IOError:
		return 0
	else:
		ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
		cur = str(_move_cursor(ws, data_in[0]))
		for i, o in zip(ws['B'+cur:'K'+cur][0], data_in[1:]):
			o.value = i
		wb.save(dst)
		return 1

def read(filename):
	''' return * in filename '''
	dst = os.path.join(FOLDER, filename)
	try:
		wb = load_workbook(dst)
	except IOError:
		print('IOError during read({})'.format(dst))
		return []
	else:
		print('load sucessully!')
		ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
		end = _move_cursor(ws)
		if end == 3:  # empty sheet
			return []
		rtr = []
		for row in tuple(map(str, range(3, end))):  # ('3', '4', '5', ..., '{end-1}')
			tmp = []  # single person container
			for col in 'ABCDEFGHIJK':
				tmp.append(ws[col+row].value)
			rtr.append(tmp)
		return rtr  # everyone contained
